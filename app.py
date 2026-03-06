"""
EBR Dashboard Backend — Polaris Bank
Uses only openpyxl — no pandas, no numpy, no compiler required.
"""

import json
import traceback
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads"
FRONT_DIR  = BASE_DIR / "frontend"
DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

app = Flask(__name__)  # No static folder — we serve HTML manually to avoid route conflicts
CORS(app)

@app.after_request
def add_no_cache_headers(response):
    """Force browsers to always fetch fresh content — no stale caches."""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def save_json(key, data):
    with open(DATA_DIR / f"{key}.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)

def load_json(key):
    p = DATA_DIR / f"{key}.json"
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return None

def flt(v):
    if v is None: return 0.0
    try:
        f = float(v)
        return 0.0 if f != f else f
    except: return 0.0

def s(v): return str(v).strip() if v is not None else ""

def ws_rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]

def find_hdr(rows, kws):
    for i, row in enumerate(rows):
        ru = [s(c).upper() for c in row]
        if any(kw.upper() in cell for cell in ru for kw in kws):
            return i
    return 0

def open_wb(fp):
    return load_workbook(fp, read_only=True, data_only=True)

# ── PROCESSORS ────────────────────────────────────────────────
def proc_opex_2024(fp):
    wb = open_wb(fp)
    sn = next((s for s in wb.sheetnames if any(k in s.upper() for k in ["2024","MOM","ACTUAL"])), wb.sheetnames[0])
    rows = ws_rows(wb[sn]); wb.close()
    h = find_hdr(rows, ["GL CODE","GL DESCRIPTION"]); hdr = rows[h]
    cc = dc = -1; mc = {}
    for ci, c in enumerate(hdr):
        cs = s(c).upper()
        if "GL CODE" in cs: cc = ci
        elif "DESCRIPTION" in cs: dc = ci
        else:
            for mi, mn in enumerate(MONTHS):
                if mn.upper() in cs and "2024" in cs: mc[mi] = ci; break
    result = {}
    for row in rows[h+1:]:
        code = s(row[cc]) if cc >= 0 and cc < len(row) else ""
        desc = s(row[dc]) if dc >= 0 and dc < len(row) else ""
        if not desc or not code or not code[:1].isdigit(): continue
        vals = [flt(row[mc[i]]) if i in mc and mc[i] < len(row) else 0.0 for i in range(12)]
        if sum(vals) == 0: continue
        result[desc] = vals
    return result

def proc_opex_2025(fp):
    wb = open_wb(fp)
    sn = next((s for s in wb.sheetnames if any(k in s.upper() for k in ["YTD","2025"])), wb.sheetnames[0])
    rows = ws_rows(wb[sn]); wb.close()
    h = find_hdr(rows, ["GL CODE","GL DESCRIPTION"]); hdr = rows[h]
    cc = dc = -1; mc = {}
    for ci, c in enumerate(hdr):
        cs = s(c).upper()
        if "GL CODE" in cs: cc = ci
        elif "DESCRIPTION" in cs: dc = ci
        else:
            for mi, mn in enumerate(MONTHS):
                if mn.upper() in cs: mc[mi] = ci; break
    result = {}
    for row in rows[h+1:]:
        code = s(row[cc]) if cc >= 0 and cc < len(row) else ""
        desc = s(row[dc]) if dc >= 0 and dc < len(row) else ""
        if not desc or not code or not code[:1].isdigit(): continue
        ytd = [flt(row[mc[i]]) if i in mc and mc[i] < len(row) else 0.0 for i in range(12)]
        mom = [ytd[0]] + [max(0.0, ytd[i]-ytd[i-1]) for i in range(1,12)]
        if sum(mom) == 0: continue
        result[desc] = mom
    return result

def proc_properties(fp):
    wb = open_wb(fp); rows = ws_rows(wb[wb.sheetnames[0]]); wb.close()
    h = find_hdr(rows, ["STATE","DESCRIPTION","TITLE","USE"]); hdr = rows[h]
    col = {}
    for ci, c in enumerate(hdr):
        cs = s(c).upper()
        if "S/N" in cs or cs in ["SN","NO"]: col["sn"] = ci
        elif "BRN" in cs or "BRANCH CODE" in cs: col["brn"] = ci
        elif "DESCRIPTION" in cs or "NAME" in cs: col["desc"] = ci
        elif "ADDRESS" in cs: col["address"] = ci
        elif "STATE" in cs: col["state"] = ci
        elif "TITLE" in cs: col["title"] = ci
        elif "USE" in cs or "TYPE" in cs: col["use"] = ci
    records = []
    for i, row in enumerate(rows[h+1:], 1):
        def g(k): return s(row[col[k]]) if k in col and col[k] < len(row) else ""
        if not g("desc") and not g("state"): continue
        records.append({"sn":g("sn") or str(i),"brn":g("brn"),"desc":g("desc"),
                        "address":g("address"),"state":g("state"),"title":g("title"),"use":g("use")})
    return records

def proc_fleet(fp):
    wb = open_wb(fp)
    sn = next((n for n in wb.sheetnames if "IN USE" in n.upper() or "FLEET" in n.upper()), wb.sheetnames[0])
    rows = ws_rows(wb[sn]); wb.close()
    records = []
    for row in rows:
        loc = None; nums = []
        for cell in row:
            v = s(cell)
            if v and v.upper() == v and len(v) > 2 and all(c.isalpha() or c in " /&-" for c in v):
                loc = v
            elif cell is not None:
                f = flt(cell)
                if f > 0: nums.append(int(f))
        if not loc or len(nums) < 3: continue
        records.append({"loc":loc,"in1016":nums[0],"in17p":nums[1],"grounded":nums[2],
                        "total":nums[3] if len(nums)>3 else nums[0]+nums[1]+nums[2]})
    return records

def proc_generators(fp):
    wb = open_wb(fp)
    sn = next((n for n in wb.sheetnames if "SUMMARY" in n.upper()), wb.sheetnames[0])
    rows = ws_rows(wb[sn]); wb.close()
    records = []
    for row in rows:
        for ci, cell in enumerate(row):
            cap = s(cell)
            if not cap or not cap[:1].isdigit(): continue
            nums = [int(flt(row[ci+j])) for j in range(1,8) if ci+j < len(row) and row[ci+j] is not None and flt(row[ci+j]) > 0]
            count = nums[0] if nums else 0
            if count == 0 or count > 500: continue
            records.append({"cap":cap,"count":count,"func":nums[1] if len(nums)>1 else 0,
                            "repair":nums[2] if len(nums)>2 else 0,"y05":nums[3] if len(nums)>3 else 0,
                            "y610":nums[4] if len(nums)>4 else 0,"y10p":nums[5] if len(nums)>5 else 0})
            break
    return records

def proc_atms(fp):
    wb = open_wb(fp)
    sn = next((n for n in wb.sheetnames if "ATM" in n.upper()), wb.sheetnames[0])
    rows = ws_rows(wb[sn]); wb.close()
    h = find_hdr(rows, ["REGION","TERMINAL","BRANCH"])
    records = []
    for i, row in enumerate(rows[h+1:], 1):
        if len(row) < 3: continue
        region = s(row[1]) if len(row) > 1 else ""
        branch = s(row[2]) if len(row) > 2 else ""
        if not region and not branch: continue
        records.append({"sn":s(row[0]) or str(i),"region":region,"branch":branch,
                        "code":s(row[3]) if len(row)>3 else "",
                        "terminal":s(row[4]) if len(row)>4 else "",
                        "id":s(row[5]) if len(row)>5 else ""})
    return records

def proc_contracts(fp):
    wb = open_wb(fp); records = []
    for sn in wb.sheetnames:
        rows = ws_rows(wb[sn])
        h = find_hdr(rows, ["LPO","DESCRIPTION","VENDOR","SUPPLIER"]); hdr = rows[h]
        col = {}
        for ci, c in enumerate(hdr):
            cs = s(c).upper()
            if "LPO NUMBER" in cs or "LPO NO" in cs: col["lpo"] = ci
            elif "DESCRIPTION" in cs or "ITEM" in cs: col["desc"] = ci
            elif "VENDOR" in cs or "SUPPLIER" in cs: col["vendor"] = ci
            elif ("TOTAL" in cs or "AMOUNT" in cs) and "total" not in col: col["total"] = ci
        for row in rows[h+1:]:
            def g(k): return s(row[col[k]]) if k in col and col[k] < len(row) else ""
            desc = g("desc")
            if not desc or len(desc) < 3: continue
            total = flt(row[col["total"]]) if "total" in col and col["total"] < len(row) else 0
            if total == 0:
                for cell in row:
                    f = flt(cell)
                    if f > 10000: total = f; break
            d = desc.upper()
            cat = ("Vehicles" if any(w in d for w in ["VEHICLE","LANDCRUISER","TOYOTA","LEXUS","HILUX","PRADO"]) else
                   "Technology" if any(w in d for w in ["LAPTOP","COMPUTER","SOFTWARE","TOKEN","SYSTEM","LICENCE","LICENSE"]) else
                   "Power & Generators" if any(w in d for w in ["GENERATOR","UPS","INVERTER","BATTERY","KVA"]) else
                   "Furniture" if any(w in d for w in ["FURNITURE","CHAIR","TABLE","CABINET"]) else
                   "Stationery" if any(w in d for w in ["STATIONERY","PAPER","PEN"]) else
                   "Renovation" if any(w in d for w in ["RENOVAT","REFURBISH","CONSTRUCTION"]) else
                   "Insurance" if "INSUR" in d else "Other")
            records.append({"month":sn,"lpo":g("lpo"),"desc":desc,"vendor":g("vendor"),"cat":cat,"total":total})
    wb.close(); return records

def proc_isg(fp):
    wb = open_wb(fp); rows = ws_rows(wb[wb.sheetnames[0]]); wb.close()
    h = find_hdr(rows, ["NAME","GENDER","REGION"]); hdr = rows[h]
    col = {}
    for ci, c in enumerate(hdr):
        cs = s(c).upper()
        if "GENDER" in cs: col["gender"] = ci
        elif "REGION" in cs: col["region"] = ci
        elif "BRANCH" in cs: col["branch"] = ci
        elif "NAME" in cs: col["name"] = ci
    male = female = 0; regions = set(); branches = set(); staff = []
    for row in rows[h+1:]:
        def g(k): return s(row[col[k]]) if k in col and col[k] < len(row) else ""
        gn = g("gender").upper()
        if gn == "M": male += 1
        elif gn == "F": female += 1
        r = g("region"); b = g("branch"); n = g("name")
        if r: regions.add(r)
        if b: branches.add(b)
        if n: staff.append({"name":n,"gender":g("gender"),"branch":b,"region":r})
    return {"total":male+female,"male":male,"female":female,"regions":len(regions),"branches":len(branches),"staff":staff}

PROCESSORS = {
    "opex2024": proc_opex_2024, "opex2025": proc_opex_2025,
    "properties": proc_properties, "fleet": proc_fleet,
    "generators": proc_generators, "atms": proc_atms,
    "contracts": proc_contracts, "isg": proc_isg,
}

# ── ROUTES ────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(str(FRONT_DIR), "dashboard.html")

@app.route("/<path:filename>")
def static_files(filename):
    """Serve frontend static files — only non-API paths reach here."""
    return send_from_directory(str(FRONT_DIR), filename)

@app.route("/api/status")
def api_status():
    out = {}
    for k in PROCESSORS:
        p = DATA_DIR / f"{k}.json"
        if p.exists():
            ts = datetime.fromtimestamp(p.stat().st_mtime).strftime("%d %b %Y %H:%M")
            out[k] = {"loaded": True, "updated": ts}
        else:
            out[k] = {"loaded": False, "updated": None}
    return jsonify({"ok": True, "status": out})

@app.route("/api/upload/<dataset>", methods=["POST"])
def api_upload(dataset):
    if dataset not in PROCESSORS: return jsonify({"ok":False,"error":"Unknown dataset"}), 400
    if "file" not in request.files: return jsonify({"ok":False,"error":"No file"}), 400
    f = request.files["file"]
    sp = UPLOAD_DIR / f"{dataset}_{f.filename}"
    f.save(str(sp))
    try:
        data = PROCESSORS[dataset](str(sp))
        save_json(dataset, data)
        count = len(data)
        return jsonify({"ok":True,"dataset":dataset,"records":count,
                        "message":f"Processed {count} records","updated":datetime.now().strftime("%d %b %Y %H:%M")})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/api/data/<dataset>")
def api_data(dataset):
    if dataset not in PROCESSORS: return jsonify({"ok":False}), 400
    d = load_json(dataset)
    if d is None: return jsonify({"ok":False,"empty":True,"error":"Not uploaded"})
    return jsonify({"ok":True,"data":d})

@app.route("/api/data/all")
def api_all():
    return jsonify({"ok":True,"data":{k:load_json(k) for k in PROCESSORS if load_json(k) is not None}})

@app.route("/api/achievements", methods=["GET"])
def get_ach(): return jsonify({"ok":True,"data":load_json("achievements") or []})

@app.route("/api/achievements", methods=["POST"])
def save_ach():
    d = request.get_json()
    save_json("achievements", d)
    return jsonify({"ok":True,"count":len(d)})

if __name__ == "__main__":
    print("\n" + "="*52)
    print("  EBR Dashboard — http://localhost:5000")
    print("="*52 + "\n")
    app.run(debug=False, port=5000, host="0.0.0.0")
